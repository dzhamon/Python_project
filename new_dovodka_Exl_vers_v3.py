# -*- coding: utf-8 -*-
"""
Created on Mon Jun 27 08:21:02 2022

@author: dzhamshed.amonov
"""

from PyQt5 import QtWidgets
from for_dovodka import Ui_MainWindow
import math
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
import pandas as pd
pd.options.display.float_format = '{:,.2f}'.format

# функция выделения уникальных элементов в массиве
def get_unique_only(st):
    #Empty list
    lst1 = []
    count = 0
    # traverse the array
    for i in st:
        if i != 0:
            if i not in lst1:
                count += 1
                lst1.append(i)
    return lst1

class mywindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(mywindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # self.ui.btn.setText("Выберите файл")

        # Здесь будут методы реакции на нажатие кнопок
        self.ui.btn_1.clicked.connect(self.choose_file) # choose_file функция действия по нажатию на кнопку

    def choose_file(self):
        path, filter = QtWidgets.QFileDialog.getOpenFileName(self, 'Select file',
            '', 'All files (*);;CSV files (*.csv)')
        self.ui.led_1.setText(path)
        new_path = path[0:len(path)-5] + '_1' + '.xlsx'
        self.ui.led_2.setText(new_path)
    
        file_string = path

        wb = load_workbook(file_string)
        ws = wb.active

        reserved_colours = [ '00FFFF',
                            '00FF00',
                            'FFFF00',
                            'FF0000',
                             'FFFACD',
                            '7FFFD4',
                            'F0E68C',
                              'B0C4DE',
                             'FFDEAD',
                             'BC8F8F',
                           'DDDDDD']

        # эти строчки дают размер обрабатываемой таблицы Excel
        rows_total = ws.max_row
        cols_total = ws.max_column

            # Цикл перебирает все строки столбца B и при ячейке равной "Общая информация",
            # выводит на печать необходимые данные и выходит досрочно из цикла
        for cell in ws['B']:
            if cell.value == 'Общая информация':
                line_num = cell.row
                col_num = cell.column
                coordinate_cell = cell.coordinate
                break
            
        contragent_names = []
        names_column = []
        for i in range(14,cols_total,1):
            if ws.cell(8,i).value != None:
                contragent_names.append(ws.cell(8,i).value)
                names_column.append(i)
            
            # поиск и фиксация Наличия столбца "Итоговая сумма со скидкой"
            
        tmp_list = []
        tmp_nums = []
            
        for k in range(14, cols_total):
            if ws.cell(9,k).value != None:
                tmp_list.append(ws.cell(9,k).value)
                tmp_nums.append(k)
                    
            """ C учетом полученных результатов выше"""
        columns_interval = []
        for i in range(len(tmp_list)-1):
            # позиции ячеек "Итоговая сумма со скидкой"
            if tmp_list[i] == 'Сумма' and tmp_list[i+1] == "Итоговая сумма со скидкой":
                columns_interval.append(tmp_nums[i+1])
                if len(columns_interval) == len(names_column):
                    break
            else:
                if tmp_list[i] == 'Сумма' and tmp_list[i+1] == "Цена за ед. в EUR":
                    columns_interval.append(0)
                if len(columns_interval) == len(names_column):
                    break
            
            # Создадим словарь из наименований исполнителей и их мест в Таблице
            
        dict_names = {}
        dict_names = dict(zip(contragent_names,names_column))
 
            # Приступаем к расчету цен на ТМЦ в валюте Евро
            
        eur_prices = np.zeros((line_num-11, len(contragent_names)), dtype='float') # создаем массив
            
            # рассчитываем суммы в Евро с учетом наличия столбца "Итоговая сумма"
            # и Тех. часть - "Соответствует", "Не соответствует"
            
        """ Рассчитываем мин суммы в единой валюте Евро"""
            
        m = 0
        k = 0
        for n in range(len(names_column)):
            k = 0
            for i in range(10, line_num - 1):
                try:
                    if columns_interval[n]!=0:
                        # при наличии столбца Сумма со скидкой
                        if ws.cell(i, names_column[n]).value=="Не соответствует":
                            eur_prices[k, m] = 0.0
                        else:
                            eur_prices[k, m] = ws.cell(i, names_column[n] + 8).value * ws.cell(i,
                            columns_interval[n] + 2).value
                    else:
                        # при отсутствии столбца Сумма со скидкой
                        if ws.cell(i, names_column[n]).value=="Не соответствует":
                            eur_prices[k, m] = 0.0
                        else:
                            eur_prices[k, m] = ws.cell(i, names_column[n] + 8).value * ws.cell(i,
                            names_column[n] + 14).value
                except TypeError:
                    pass
                k = k + 1
            m = m + 1
        # формирование списка ненулевых минимальных элементов
        # по строкам массива eur_prices
            
        min_values = []
        for i in range(eur_prices.shape[0]):
            s = eur_prices[i,:]
            min_values.append(np.nanmin(np.where(s == 0, np.nan, s)))
            
        # В этом двойном цикле выбираем по минимальным значениям цен в Евро
        #коммерческие предложения потенциальных контрагентов-победителей
            
        position_number = []
        for i in range(eur_prices.shape[0]):
            for j in range(eur_prices.shape[1]):
                if min_values[i] == eur_prices[i,j]:
                    pos_number = j
                    # теперь по номеру позиции находим потенциального поставщика
                    position_number.append(dict_names.get(contragent_names[pos_number])) # выдергивает значение по ключу из dict_names
            
            # выделим уникальные номера потенциальных поставщиков и присвоим им цвета (создав словарь)
        unique_position_numbers = get_unique_only(position_number)
            # здесь определяются идентификаторы суммы поставок от потенциального поставщика
        sum_positions = []
        for unique_position_number in unique_position_numbers:
            ss = 'sum_' + str(unique_position_number)
            sum_positions.append(ss)
            
        for i in range(len(sum_positions)):
            sum_positions[i] = 0.0
            
        # это словарь: key - номера позиций поставщиков, values - суммы их потенциальных поставок
        dict_positions = {}
        dict_positions = dict(zip(unique_position_numbers, sum_positions))
        
        # далее словарь, с данными по позициям и выделенным им цветам
        position_colors = []
        for i in range(len(unique_position_numbers)):
            position_colors.append(reserved_colours[i])
        
        dict_temp_colors = {}
        dict_temp_colors = dict(zip(unique_position_numbers,position_colors))
            
        # Далее. Формируем выходную таблицу и в ней занимаемся отбором поставок и
        # выделение их цветом
            
        ws.insert_cols(cols_total,len(contragent_names)+3) # раздвигаем таблицу и вставляем столбцы
            
        # записываем имена контрагентов в вновь созданные столбцы
        j = 0
        for i in range(cols_total+1, cols_total+1+len(contragent_names),1):
            ws.cell(row=9, column=i, value= contragent_names[j])
            j += 1
            
        tmp_cell = ws.cell(row=9, column=i+2, value='MIN EUR')
        tmp_cell.fill = PatternFill('solid', fgColor="DDDDDD")
        pos_min_eur = i+2
            
        # этот блок записывает в новые ячейки массив eur_prices всех контрагентов и
        # столбец с минимальными значениями цен в Eur (min_value)
        pos_num_col = cols_total+1
        for j in range(eur_prices.shape[1]):
            pos_num_row = 10
            for i in range(eur_prices.shape[0]):
                ws.cell(row=pos_num_row, column=pos_num_col, value = eur_prices[i,j])
                pos_num_row += 1
            pos_num_col += 1
            
            # теперь распечатаем найденные мин значения ценовых предложений
            
        for i in range(len(min_values)):
            ws.cell(row=10+i, column=pos_num_col+1, value=min_values[i])
            
        # Создадим массив для записи миним цен потенциальных победителей
            
        potential_prices = np.zeros((eur_prices.shape[0], len(dict_temp_colors)), dtype='float')
            
            #  выделим цветами конкурентные цены в Евро потенциальных победителей
            
        pos_num_row = 9
        pos_num_col = cols_total + 1
        
        k = 0
        winners_num = []
        for i in range(pos_num_col, pos_num_col+(eur_prices.shape[1])):
            stolbec = pos_num_col + k
            var_1 = dict_names.get(ws.cell(row=pos_num_row, column=stolbec).value)
            var_2 = dict_temp_colors.get(var_1, 0)
            if var_2 != 0:
                colored_cell = ws.cell(row=pos_num_row, column=stolbec)
                colored_cell.fill = PatternFill('solid', fgColor=var_2)
                winners_num.append(stolbec)
            k = k + 1
            
        """ Выделим в Таблице отведенными цветами предложения победителей """
            
        # загоним мин значения предложений победителей в массив potential_prices
        i = 0
        for winner_num in winners_num:
            for j in range(len(min_values)):
                potential_prices[j,i] = ws.cell(row=pos_num_row+1+j,column=winner_num).value
            i = i + 1
            
            # приступаем непосредственно к покраске строк
            # одновременно подсчитываются общие стоимости поставки каждого победителя
            # и записываются в словаре dict_positions
            
        idx_temp_list = []
        i = 0
        for i in range(len(min_values)):
            idx_temp_list.append(i)
            
        dict_prices = dict(zip(min_values, idx_temp_list))
        
        k = 0
        for m in range(len(min_values)):    # min_value in min_values
            if math.isnan(min_values[m]):
                k = k + 1
                m = m + 1
            else:
               for j in range(eur_prices.shape[1]):
                    if min_values[m] == eur_prices[k,j]:
                        column_tmp = pos_num_col + j
                        winn_name = ws.cell(row=pos_num_row, column=column_tmp).value
                        value_position = dict_names.get(winn_name)
                        color_num = dict_temp_colors.get(value_position)
                        for i in range(value_position, value_position+16,1):
                            tmp_cell = ws.cell(row=pos_num_row+1+k, column=i)
                            tmp_cell.fill = PatternFill('solid', fgColor=color_num)
                        ss_temp = dict_positions.get(value_position)
                        temp_s = ws.cell(row=pos_num_row+1+k, column=value_position + 12).value
                        ss_temp = ss_temp + temp_s
                        dict_positions[value_position] = ss_temp
                        k = k + 1
                        break
                    else:
                        continue
            
            # разместим полученные результаты в соответствующих ячейках таблицы
            # добавим строку
            
        ws.insert_rows(line_num, amount=1)
            
        for position in dict_positions.keys():
            pos = position
            ws.cell(row= line_num, column=pos).value = dict_positions.get(position)
            
        wb.save(new_path)
        self.ui.led_3.setText('Файл сформирован!')
        self.ui.led_3.setStyleSheet("color: rgb(247,0,0);")

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication([])
    application = mywindow()
    application.show()
    sys.exit(app.exec_())
    
    